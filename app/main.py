import json
import os
import sys
import uuid
from pathlib import Path
from typing import Optional

import tempfile
from io import BytesIO

import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.calculator import calculate_strategy
from app.model_service import predict, predict_batch
from app.schemas import BatchClientData, ClientData, StrategyInput

app = FastAPI(title='Bank Marketing - Система прогнозирования депозитов')

static_dir = Path(__file__).resolve().parent.parent / 'static'
static_dir.mkdir(exist_ok=True)
app.mount('/static', StaticFiles(directory=str(static_dir)), name='static')

templates_dir = Path(__file__).resolve().parent / 'templates'
templates_dir.mkdir(exist_ok=True)
templates = Jinja2Templates(directory=str(templates_dir))


def load_json(path):
    path = Path(__file__).resolve().parent.parent / path
    if path.exists():
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def get_charts():
    path = Path(__file__).resolve().parent.parent / 'charts_manifest.json'
    if path.exists():
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    charts_dir = Path(__file__).resolve().parent.parent / 'static' / 'charts'
    charts = {}
    if charts_dir.exists():
        for fname in sorted(charts_dir.glob('*.png')):
            charts[fname.stem] = str(fname.relative_to(
                Path(__file__).resolve().parent.parent))
    return charts


@app.get('/', response_class=HTMLResponse)
async def index(request: Request):
    analysis = load_json('analysis_results.json')
    recommendations = load_json('recommendations.json')
    charts = get_charts()
    return templates.TemplateResponse('index.html', {
        'request': request,
        'analysis': analysis,
        'recommendations': recommendations,
        'charts': charts,
    })


@app.get('/api/analysis')
async def get_analysis():
    return {
        'analysis': load_json('analysis_results.json'),
        'recommendations': load_json('recommendations.json'),
        'charts': get_charts(),
    }


@app.post('/api/predict', response_model=dict)
async def predict_single(client: ClientData):
    try:
        result = predict(client.model_dump())
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/api/predict/batch', response_model=dict)
async def predict_batch_api(data: BatchClientData):
    try:
        clients_dict = [c.model_dump() for c in data.clients]
        result = predict_batch(clients_dict)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/api/strategy', response_model=dict)
async def get_strategy(params: StrategyInput):
    try:
        result = calculate_strategy(
            cost_per_call=params.cost_per_call,
            profit_per_deposit=params.profit_per_deposit,
            budget=params.budget,
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# In-memory store for strategy Excel downloads
_strategy_downloads = {}


@app.post('/api/strategy/upload')
async def strategy_upload(
    file: UploadFile = File(...),
    cost_per_call: float = Form(...),
    profit_per_deposit: float = Form(...),
    budget: Optional[float] = Form(None),
):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='Файл должен быть в формате .xlsx')
    try:
        content = await file.read()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_in:
            tmp_in.write(content)
            tmp_in_path = tmp_in.name
        try:
            df_input = pd.read_excel(tmp_in_path, engine='openpyxl')
            df = df_input.rename(columns=COLUMN_MAP_UPLOAD)
            missing = [c for c in FEATURE_ORDER_UPLOAD if c not in df.columns]
            if missing:
                raise HTTPException(
                    status_code=400,
                    detail=f'В файле отсутствуют столбцы: {missing}.')
            for col in ['default', 'housing', 'loan']:
                df[col] = df[col].map(lambda x: YES_NO_MAP.get(str(x).lower(), 0))

            from app.model_service import get_model
            model = get_model()
            probabilities = model.predict_proba(df[FEATURE_ORDER_UPLOAD])[:, 1]
            df['probability'] = probabilities

            result = calculate_strategy(
                cost_per_call=cost_per_call,
                profit_per_deposit=profit_per_deposit,
                budget=budget,
                df=df,
            )

            # Build sorted Excel with cutoff marker
            df_sorted = df_input.copy()
            df_sorted['Вероятность, %'] = (probabilities * 100).round(2)
            df_sorted = df_sorted.sort_values('Вероятность, %', ascending=False).reset_index(drop=True)
            df_sorted.index = df_sorted.index + 1

            cutoff_row = result['clients_to_call']
            df_sorted['_gap1'] = ''
            df_sorted['Обзванивать'] = ['да' if i <= cutoff_row else 'нет' for i in df_sorted.index]
            df_sorted['_gap2'] = ''
            # Reorder: original cols, gap, prob, gap, call, gap
            orig_cols = [c for c in df_input.columns]
            new_order = orig_cols + ['_gap1', 'Вероятность, %', 'Обзванивать', '_gap2']
            df_sorted = df_sorted[new_order]

            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            token = str(uuid.uuid4())
            tmp_dir = Path(tempfile.gettempdir()) / 'strategy_downloads'
            tmp_dir.mkdir(exist_ok=True)
            out_path = tmp_dir / f'{token}.xlsx'
            df_sorted.to_excel(str(out_path), index=False, engine='openpyxl')
            wb = load_workbook(str(out_path))
            ws = wb.active
            # Clear gap column headers (col 16 and 19)
            ws.cell(row=1, column=16).value = ''
            ws.cell(row=1, column=19).value = ''
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            bold_font = Font(bold=True)
            # Highlight cutoff row
            cutoff_excel_row = cutoff_row + 1  # +1 for header
            if cutoff_excel_row <= ws.max_row:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=cutoff_excel_row, column=col).fill = green_fill
                    ws.cell(row=cutoff_excel_row, column=col).font = bold_font
            # Add note row after cutoff
            note_row = cutoff_excel_row + 1
            ws.insert_rows(note_row, 1)
            ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ws.max_column)
            ws.cell(row=note_row, column=1).value = f'↑ Обзванивать до этой строки ({cutoff_row} клиентов)'
            ws.cell(row=note_row, column=1).font = Font(bold=True, color='006100')
            ws.cell(row=note_row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=note_row, column=1).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            wb.save(str(out_path))
            wb.close()
            _strategy_downloads[token] = str(out_path)

            result['download_url'] = f'/api/strategy/download/{token}'
            result['total_clients'] = len(df_sorted)
            return result

        finally:
            os.unlink(tmp_in_path)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Ошибка: {str(e)}')


@app.get('/api/strategy/download/{token}')
async def strategy_download(token: str):
    path = _strategy_downloads.get(token)
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail='Файл не найден')
    with open(path, 'rb') as f:
        data = f.read()
    os.unlink(path)
    _strategy_downloads.pop(token, None)
    return StreamingResponse(
        BytesIO(data),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=strategy_result.xlsx'}
    )


COLUMN_MAP_UPLOAD = {
    'Возраст': 'age', 'Работа': 'job', 'Семейное положение': 'marital',
    'Образование': 'education', 'Дефолт': 'default',
    'Среднегодовой баланс, в евро': 'balance',
    'Наличие жилищного кредита': 'housing',
    'Наличие потребительского кредита': 'loan',
    'Тип контактной связи ': 'contact',
    'Последний контакт, месяц': 'month',
    'Длительность контакта, секунд': 'duration',
    'Количество контактов, выполненных во время этой кампании и для данного клиента (числовое, включая последний контакт)': 'campaign',
    'Количество дней, прошедших с момента последнего контакта с клиентом из предыдущей кампании (число, -1 означает, что с клиентом ранее не связывались)': 'pdays',
    'Количество контактов, выполненных до этой кампании и для этого клиента': 'previous',
    'Результат предыдущей маркетинговой кампании': 'poutcome',
}
COLUMN_MAP_REVERSE = {v: k for k, v in COLUMN_MAP_UPLOAD.items()}
FEATURE_ORDER_UPLOAD = ['age', 'job', 'marital', 'education', 'default', 'balance',
                        'housing', 'loan', 'contact', 'month', 'duration', 'campaign',
                        'pdays', 'previous', 'poutcome']
YES_NO_MAP = {'да': 1, 'нет': 0}


@app.post('/api/predict/upload')
async def predict_upload(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='Файл должен быть в формате .xlsx')
    try:
        content = await file.read()

        # Use temp file for Excel processing (avoids encoding issues with BytesIO)
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_in:
            tmp_in.write(content)
            tmp_in_path = tmp_in.name

        try:
            df_input = pd.read_excel(tmp_in_path, engine='openpyxl')

            # Map columns
            df = df_input.rename(columns=COLUMN_MAP_UPLOAD)
            missing = [c for c in FEATURE_ORDER_UPLOAD if c not in df.columns]
            if missing:
                raise HTTPException(
                    status_code=400,
                    detail=f'В файле отсутствуют столбцы: {missing}. '
                           f'Убедитесь, что названия столбцов совпадают с исходным файлом данные.xlsx')

            # Encode binary features
            for col in ['default', 'housing', 'loan']:
                df[col] = df[col].map(lambda x: YES_NO_MAP.get(str(x).lower(), 0))

            # Predict probabilities
            from app.model_service import get_model
            model = get_model()
            probabilities = model.predict_proba(df[FEATURE_ORDER_UPLOAD])[:, 1]

            # Build result with ASCII column names only
            df_result = df_input.copy()
            df_result['Вероятность, %'] = (probabilities * 100).round(2)
            df_result = df_result.sort_values('Вероятность, %', ascending=False).reset_index(drop=True)
            df_result.index = df_result.index + 1

            # Write to temp file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_out:
                df_result.to_excel(tmp_out.name, index=False, engine='openpyxl')
                tmp_out_path = tmp_out.name

            with open(tmp_out_path, 'rb') as f:
                output = BytesIO(f.read())
            os.unlink(tmp_out_path)
        finally:
            os.unlink(tmp_in_path)

        output.seek(0)
        safe_name = file.filename.replace('.xlsx', '')
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={safe_name}_sorted.xlsx'}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Ошибка обработки: {str(e)}')


@app.get('/api/health')
async def health():
    return {'status': 'ok', 'model_loaded': os.path.exists('model/catboost_model.cbm')}


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host='0.0.0.0', port=8000)
